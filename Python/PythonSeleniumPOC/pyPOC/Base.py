import openpyxl
from selenium.webdriver import Chrome, ActionChains
from datetime import datetime

path = ("E:/seleniumDriver/browserdriver/chromedriver.exe")
driver = Chrome(executable_path=path)
driver.get("http://automationpractice.com/index.php")
driver.maximize_window()
print(driver.title)
driver.get_screenshot_as_file("website.png")
driver.implicitly_wait(5)
excelPath = "E:\\POC2.xlsx"
book_path = openpyxl.load_workbook(excelPath)
sheet_obj = book_path.active
l_row = sheet_obj.max_row
l_col = sheet_obj.max_column
Title = driver.title
assert Title == "My Store"
url = driver.current_url
print(url)
now = datetime.now()
Date_time = now.strftime("%Y%m%d%H%M%S")
string_date = str(Date_time)
s1_login = ("Login" + string_date + ".png")
S2_product = ("product" + string_date + ".png")
S3_Add = ("add" + string_date + ".png")
S4_Proceed = ("proceed" + string_date + ".png")
S5_check_out = ("checkout" + string_date + ".png")

for i in range(2, l_row + 1):
    Login = driver.find_element_by_xpath("//a[@class='login']")
    Login.click()
    Username_text = driver.find_element_by_css_selector("#email")
    Username_text.send_keys(sheet_obj.cell(i, column=1).value)
    password = driver.find_element_by_css_selector("#passwd")
    password.send_keys(sheet_obj.cell(i, column=2).value)
    driver.get_screenshot_as_file(s1_login)
    Sub = driver.find_element_by_css_selector("#SubmitLogin")
    Sub.click()
    action = ActionChains(driver)
    select_tab = driver.find_element_by_xpath(
        "//ul[@class='sf-menu clearfix menu-content sf-js-enabled sf-arrows']/li[1]")
    action.move_to_element(select_tab).perform()
    option = driver.find_element_by_xpath(
        "//ul[@class='submenu-container clearfix first-in-line-xs']//a[text()='T-shirts']")
    option.click()
    heading = driver.find_element_by_xpath("//span[@class='category-name']")
    heading_text = heading.text
    assert heading_text == "T-shirts"
    driver.execute_script("window.scrollTo(0, 800)")
    product = driver.find_element_by_xpath("//span[@class='available-now']")
    product.click()
    driver.get_screenshot_as_file(S2_product)
    dropdown = driver.find_element_by_xpath("//a[@class='quick-view']")
    More = driver.find_element_by_xpath("//span[text()='More']")
    More.click()
    item_title = driver.find_element_by_xpath("//h1[@itemprop='name']").text
    assert item_title == "Faded Short Sleeve T-shirts"
    add = driver.find_element_by_xpath("//span/i[@class='icon-plus']")
    add.click()
    driver.get_screenshot_as_file(S3_Add)
    drop_down_list = driver.find_elements_by_xpath("//select[@id='group_1']/option")
    len_dropdown = len(drop_down_list)
    for len_dropdown in drop_down_list:
        if len_dropdown.text == 'M':
            len_dropdown.click()
            break
    driver.execute_script("window.scrollTo(0, 200)")
    color = driver.find_element_by_xpath("//a[@name='Orange']")
    color.click()
    submit = driver.find_element_by_xpath("//span[text()='Add to cart']")
    submit.click()
    driver.get_screenshot_as_file(S4_Proceed)
    proceed = driver.find_element_by_xpath("//a[@title='Proceed to checkout']")
    proceed.click()
    assert_check = driver.find_element_by_css_selector(".navigation_page").text
    assert assert_check == "Your shopping cart"
    driver.execute_script("window.scrollTo(0, 200)")
    checkout = driver.find_element_by_xpath("//span[text()='Proceed to checkout']")
    checkout.click()
    driver.get_screenshot_as_file(S5_check_out)
    Sign_out = driver.find_element_by_xpath("//a[@class='logout']")
    Sign_out.click()
    assert_sign_in = driver.find_element_by_css_selector(".login").text
    assert assert_sign_in == "Sign in"
    driver.get("http://automationpractice.com/index.php")
